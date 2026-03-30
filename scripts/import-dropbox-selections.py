"""
Import plumbing and lighting selections from the Dropbox xlsx spreadsheet
into the Supabase selections table.

Reads the "Plumbing Allowance Newp Brass" and "Lighting Allowance" sheets,
parses items with product names, prices, model numbers, and URLs,
then upserts into selections (updating existing matches, creating new ones).

Usage: python scripts/import-dropbox-selections.py
"""

import os
import json
import re
import openpyxl
from supabase import create_client

# Config
XLSX_PATH = "C:/Users/danie/Dropbox/Properties/Austin, TX/Liberty Hill/708 Purple Salvia Cove/Development/Bids/Ballpark_Estimate_10_14_25_v1_NB_Master_Shower_Filled (version 2).xlsx"

SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
SUPABASE_KEY = os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    # Try loading from .env.local
    with open(".env.local") as f:
        for line in f:
            line = line.strip()
            if line.startswith("NEXT_PUBLIC_SUPABASE_URL="):
                SUPABASE_URL = line.split("=", 1)[1]
            elif line.startswith("NEXT_PUBLIC_SUPABASE_ANON_KEY="):
                SUPABASE_KEY = line.split("=", 1)[1]

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# Get project ID
project = supabase.table("projects").select("id").limit(1).execute()
PROJECT_ID = project.data[0]["id"]
print(f"Project: {PROJECT_ID}")


def normalize_room(room_str):
    """Normalize room names to match existing DB conventions."""
    if not room_str:
        return "Whole House"
    room = str(room_str).strip()
    mappings = {
        "Kitchen/Dining/Foyer": "Kitchen",
        "Main Kitchen": "Kitchen",
        "Pantry Kitchen": "Pantry",
        "Master Bath": "Master Bath",
        "Master": "Master Bedroom",
        "Bath 2": "Bath 2",
        "Bath # 3 (Gym)": "Bath 3",
        "Bath #3 (Gym)": "Bath 3",
        "Bath 3 & 4": "Bath 3",  # Will duplicate for Bath 4
        "Bath #4 (Kids)": "Bath 4",
        "Powder": "Powder Bath",
        "Powder Room": "Powder Bath",
        "Classroom": "Classroom",
        "School Room": "Classroom",
        "Utility": "Utility / Laundry",
        "Laundry": "Utility / Laundry",
        "Laundry Room": "Utility / Laundry",
        "Dog Wash": "Utility / Laundry",
        "Garage": "Garage",
        "Great Room / Dining Room": "Great Room",
        "Bedroom 3": "Bedroom 3",
        "Bedroom 4 (Kids)": "Bedroom 4",
        "Housewide": "Whole House",
        "Exterior": "Exterior",
    }
    return mappings.get(room, room)


def parse_plumbing_sheet(wb):
    """Parse the Plumbing Allowance Newp Brass sheet.

    Column layout (0-indexed):
    - Col 1: Room header (when no col 4 data) OR fixture type
    - Col 4: Quantity
    - Col 8: Total cost (actual)
    - Col 9: Per unit cost
    - Col 10: Fixture name + model
    - Col 11: URL to buy
    """
    ws = wb["Plumbing Allowance Newp Brass"]
    items = []
    current_room = None

    skip_labels = {"Quantity", "Plumbing Fixtures", "Project Name:", "Project Address:",
                   "Unnamed: 0", "Unnamed: 1"}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]

        # Col 1 text with no col 4 (quantity) = room header
        col1 = str(vals[1]).strip() if vals[1] else None
        col4 = vals[4] if len(vals) > 4 else None

        if col1 and not col4:
            if col1 not in skip_labels and len(col1) > 2:
                current_room = col1
            continue

        # Skip if no fixture name in col 10
        col10 = str(vals[10]).strip() if len(vals) > 10 and vals[10] else None
        if not col10 or col10.startswith("Unnamed") or col10 == "Fixture name and model or SKU":
            continue

        fixture_type = str(vals[1]).strip() if vals[1] else None
        qty = int(vals[4]) if len(vals) > 4 and vals[4] and isinstance(vals[4], (int, float)) else 1
        unit_price = float(vals[9]) if len(vals) > 9 and vals[9] and isinstance(vals[9], (int, float)) else None
        url = str(vals[11]).strip() if len(vals) > 11 and vals[11] and str(vals[11]).startswith("http") else None

        # Also check hyperlinks
        if not url:
            for cell in row:
                if cell.hyperlink:
                    url = cell.hyperlink.target
                    break

        # Build product name: combine fixture name from col 10
        # Sometimes col 10 has the full name, sometimes it spans to next row
        fixture_name = col10

        # Extract model number
        model_match = re.search(r'(\d{3,5}[-/]\d{3,5}[A-Z/]*|\b[A-Z]{2,}\d{3,}[A-Z]*\b)', fixture_name)
        model = model_match.group(0) if model_match else None

        # Extract brand
        brand = None
        for b in ["Newport Brass", "TOTO", "Franke", "Aquatica", "Kingston Brass", "Kohler"]:
            if b.lower() in fixture_name.lower():
                brand = b
                break

        if current_room and fixture_name:
            items.append({
                "room": normalize_room(current_room),
                "category": "plumbing",
                "product_name": fixture_name[:200],
                "brand": brand,
                "model_number": model,
                "quantity": qty,
                "unit_price": unit_price,
                "total_price": (unit_price * qty) if unit_price else None,
                "product_url": url,
                "status": "selected",
                "source": "dropbox_import",
            })

    return items


def parse_lighting_sheet(wb):
    """Parse the Lighting Allowance sheet."""
    ws = wb["Lighting Allowance"]
    items = []
    current_room = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]

        # Room headers: col A has room name, col B is empty or "Type"
        if vals[0] and isinstance(vals[0], str) and vals[0].strip() and not vals[2]:
            room_text = vals[0].strip()
            if room_text not in ("Exterior",) and not vals[1]:
                current_room = room_text
                continue
            elif room_text == "Exterior":
                current_room = "Exterior"
                continue

        # Data rows have: type in col 1, description in col 2, qty in col 3, price in col 4, total in col 5, model in col 6, link in col 7
        description = vals[2] if len(vals) > 2 else None
        if not description or not isinstance(description, str) or len(description) < 3:
            continue

        qty = int(vals[3]) if len(vals) > 3 and vals[3] and isinstance(vals[3], (int, float)) else 1
        unit_price = float(vals[4]) if len(vals) > 4 and vals[4] and isinstance(vals[4], (int, float)) else None
        model_info = str(vals[6]).strip() if len(vals) > 6 and vals[6] else None
        url = str(vals[7]).strip() if len(vals) > 7 and vals[7] and str(vals[7]).startswith("http") else None

        # Also check for hyperlinks
        for cell in row:
            if cell.hyperlink and not url:
                url = cell.hyperlink.target

        fixture_type = str(vals[1]).strip() if vals[1] else "Fixture"

        # Extract brand from model info or description
        brand = None
        for b in ["RH", "Savoy House", "Kichler", "Visual Comfort", "Progress Lighting", "Alora",
                   "Hudson Valley", "Lantern and Scroll", "deVOL", "Birch Lane"]:
            text = (model_info or "") + " " + description
            if b.lower() in text.lower():
                brand = b
                break

        items.append({
            "room": normalize_room(current_room),
            "location_detail": fixture_type,
            "category": "lighting",
            "product_name": description[:200],
            "brand": brand,
            "model_number": model_info[:100] if model_info else None,
            "quantity": qty,
            "unit_price": unit_price,
            "total_price": (unit_price * qty) if unit_price else None,
            "product_url": url,
            "status": "selected" if unit_price else "considering",
            "source": "dropbox_import",
        })

    return items


def upsert_selections(items, category_label):
    """Upsert items into selections table, matching by product_name + room."""
    if not items:
        print(f"  {category_label}: 0 items to process")
        return 0, 0

    # Get existing selections for this category
    existing = supabase.table("selections").select("id, product_name, room, product_url, model_number").eq("project_id", PROJECT_ID).eq("category", items[0]["category"]).execute()
    existing_map = {}
    for s in existing.data:
        key = (s["product_name"].lower()[:50] if s["product_name"] else "", (s["room"] or "").lower())
        existing_map[key] = s

    created = 0
    updated = 0
    skipped = 0

    for item in items:
        key = (item["product_name"].lower()[:50], (item["room"] or "").lower())

        # Try to find a match
        match = existing_map.get(key)

        # Also try fuzzy match on product name (first 30 chars)
        if not match:
            short_key = (item["product_name"].lower()[:30], (item["room"] or "").lower())
            for ek, ev in existing_map.items():
                if ek[0][:30] == short_key[0] and ek[1] == short_key[1]:
                    match = ev
                    break

        if match:
            # Update existing with URL and any missing fields
            updates = {}
            if item.get("product_url") and not match.get("product_url"):
                updates["product_url"] = item["product_url"]
            if item.get("model_number") and not match.get("model_number"):
                updates["model_number"] = item["model_number"]
            if item.get("unit_price"):
                updates["unit_price"] = item["unit_price"]
                updates["total_price"] = item.get("total_price")
            if item.get("brand"):
                updates["brand"] = item["brand"]

            if updates:
                supabase.table("selections").update(updates).eq("id", match["id"]).execute()
                updated += 1
            else:
                skipped += 1
        else:
            # Create new selection
            record = {
                "project_id": PROJECT_ID,
                "room": item["room"],
                "location_detail": item.get("location_detail"),
                "category": item["category"],
                "product_name": item["product_name"],
                "brand": item.get("brand"),
                "model_number": item.get("model_number"),
                "quantity": item["quantity"],
                "unit_price": item.get("unit_price"),
                "total_price": item.get("total_price"),
                "product_url": item.get("product_url"),
                "status": item.get("status", "selected"),
            }
            # Remove None values
            record = {k: v for k, v in record.items() if v is not None}
            try:
                supabase.table("selections").insert(record).execute()
                created += 1
            except Exception as e:
                print(f"  ERROR creating {item['product_name'][:40]}: {e}")

    print(f"  {category_label}: {created} created, {updated} updated, {skipped} unchanged")
    return created, updated


if __name__ == "__main__":
    print(f"Reading: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print("\nParsing Plumbing Allowance Newp Brass...")
    plumbing_items = parse_plumbing_sheet(wb)
    print(f"  Found {len(plumbing_items)} plumbing items")

    print("\nParsing Lighting Allowance...")
    lighting_items = parse_lighting_sheet(wb)
    print(f"  Found {len(lighting_items)} lighting items")

    print("\nUpserting plumbing selections...")
    p_created, p_updated = upsert_selections(plumbing_items, "Plumbing")

    print("\nUpserting lighting selections...")
    l_created, l_updated = upsert_selections(lighting_items, "Lighting")

    print(f"\nDone! Plumbing: {p_created} new + {p_updated} updated. Lighting: {l_created} new + {l_updated} updated.")
